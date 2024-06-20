import glob
import zipfile
import os
from PIL import Image
import os
import openpyxl
import csv
import argparse

# to avoid DecompressionBombWarining
Image.MAX_IMAGE_PIXELS = 1000000000

img_id = 0
task_count = 0
crowd_img_count = 0
save_img_num = 0
class_names = []
all_header = []
species_names = []

root_dir = os.path.dirname(__file__)

def get_arg_parse():
    parser = argparse.ArgumentParser()

    parser.add_argument('--cvat', '-c', type=str, default='cvat_40_202404') # cvat files dir
    parser.add_argument('--save', '-s', type=str, default='debug') # save dir
    parser.add_argument('--mag', '-x', type=str, default='40') # cvat data's magnification (x40 or x100) 

    args = parser.parse_args()
    return args


args = get_arg_parse()

# path to cvat zipfiles
task_zip_paths = glob.glob(os.path.join(root_dir, args.cvat, '*.zip'))
task_zip_paths = sorted(task_zip_paths)

# path to save dir
save_dir_path = os.path.join(root_dir, args.save)


# make save dir
if not os.path.exists(save_dir_path):
    os.makedirs(save_dir_path)
# make header for anno_list
with open(save_dir_path + '/anno_list_' + args.mag + '.csv', mode='w',newline='',encoding='utf-8') as f:
    writer = csv.writer(f)
    header = ['img_id','class','order','family','genus','species','subsp','note','crowd_img','cx','cy','rw','rh','point','採集地','採集日','task_name']
    writer.writerow(header)


# main
for task_zip_path in task_zip_paths:
    with zipfile.ZipFile(task_zip_path) as zf:
        task_dir = task_zip_path.replace('.zip', '')
        zf.extractall(task_dir)

        obj_dir = os.path.join(task_dir, 'obj_train_data')
        with open(os.path.join(task_dir, 'obj.names'), 'r') as f:
            id2name = []
            subsps = []
            remarks = []

            # Since the anno_name and its order differ for each zip, id2name must be created each time.
            for line in f.readlines():
                anno = line.strip()
                anno_list = [name.strip() for name in anno.split(',')]
                anno_name = ','.join(anno_list)
                wb = openpyxl.load_workbook(os.path.join(root_dir, 'anno_name_list_202405.xlsx')) # path to anno_name list 
                # wb = openpyxl.load_workbook(os.path.join(root_dir, 'anno_name_list.xlsx')) # path to anno_name list 
                sheet = wb['Sheet1']
                first_row = sheet['A']
                column = ['C','D','E','F','G','H','I']
                stage = ['cl.','or.','fa.','ge.','sp.']  # stage = ['class','order','family','genus','species']
                name_level = []
                count = 0
                class_name = 'a'
                unknown_annos = ['?','？']
                subsp = None
                remark = None
                name_get_flag = 0

                for cell in first_row:
                    cell_anno_list = [name.strip() for name in cell.value.split(',')]
                    cell_anno_name = ','.join(cell_anno_list)
                    if cell_anno_name.lower() == anno_name.lower() and name_get_flag == 0:
                        row_number = cell.row
                        unknown_flag = 0
                        for i in range(5):
                            if unknown_flag == 1:
                                if i == 4:  # if anno_name contains a growth process or sex, add a suffix to distinguish
                                    growth = ['male','female','juvenile'] 
                                    for name in growth:
                                        anno_growth_name = anno_name.split(',')[-1].strip()
                                        if name == anno_growth_name.lower():
                                            remark = name
                                    name_level.append(class_name.lower())

                                else:
                                    name_level.append(name_level[-1])

                            else:   # debug
                                if anno_name in unknown_annos:
                                    class_name = 'All_unknown'
                                    unknown_flag += 1

                                else:
                                    class_name = sheet[column[i] + str(row_number)].value

                                    if class_name == None  and sheet[column[-2] + str(row_number)].value != None:   #class_name unknown,with stage
                                        class_name = stage[i] + '_unk_' + sheet[column[-2] + str(row_number)].value + '_stage'
                                        unknown_flag += 1
                                    elif class_name == None and i!= 0:    # class_name unknown,without stage
                                        class_name = stage[i] + '_unk'
                                        unknown_flag += 1
                                    elif i == 4 and sheet[column[-2] + str(row_number)].value != None:  # class_name known,with stage 
                                        class_name = class_name + '_' + sheet[column[-2] + str(row_number)].value + '_stage'
                                    elif i == 4 and sheet[column[-1] + str(row_number)].value != None:  # class_name known,with subsp
                                        subsp = sheet[column[-1] + str(row_number)].value
                                        

                                    if i == 4:
                                        # if anno_name contains a growth process or sex, add a suffix to distinguish
                                        growth = ['male','female','juvenile']
                                        for name in growth:
                                            anno_growth_name = anno_name.split(',')[-1].strip()
                                            if name == anno_growth_name.lower():
                                                remark = name
                                   
                                name_level.append(class_name.lower().strip())


                        class_name = ','.join(name_level)
                        all_header.append([n.capitalize() for n in name_level])
                        name_get_flag += 1

                
                if class_name == 'a': #debug
                    import pdb;
                    pdb.set_trace()
                class_name = ','.join([n.capitalize() for n in class_name.split(',')])
     
                id2name.append(class_name)
                class_names.append(class_name)
                subsps.append(subsp) 
                remarks.append(remark) 


        task_data, path_data,label_data,point_data,crop_data,point_data,id_data = [],[],[],[],[],[],[]
        subsp_data,remark_data = [],[]
        point_name = []
        data= []



        img_paths = glob.glob(os.path.join(obj_dir, '*.jpg'))    
        for img_path in img_paths:
            img = Image.open(img_path)
            width, height = img.size

            crowd_img_count += 1

            with open(img_path[: -4] + '.txt', 'r') as f:
                labels = f.readlines()
            for i, label in enumerate(labels):
                class_id, cx, cy, rw, rh = label.split(' ')
                cx, cy, rw, rh = float(cx), float(cy), float(rw), float(rh)
                left, upper, right, lower = int((cx-rw/2)*width), int((cy-rh/2)*height), int((cx+rw/2)*width), int((cy+rh/2)*height)
                img_crop = img.crop((left, upper, right, lower))

                cropped_img_dir = os.path.join(save_dir_path,'images_' + args.mag, id2name[int(class_id)])
                
                path_data.append([img_path.split('/')[-1]])
                label_data.append(id2name[int(class_id)].split(','))
                crop_data.append([cx,cy,rw,rh])
                id_data.append([f'{img_id:06d}'])
                subsp_data.append([subsps[int(class_id)]])
                remark_data.append([remarks[int(class_id)]])


                # when saving all cropped img
                os.makedirs(cropped_img_dir, exist_ok=True)
                img_crop.save(os.path.join(cropped_img_dir, f'{img_id:06d}'+'.jpg'))

                # when saving only All_unknown img (debug)
                # if id2name[int(class_id)].split(',')[0] == 'All_unknown':
                #     os.makedirs(cropped_img_dir, exist_ok=True)
                #     img_crop.save(os.path.join(cropped_img_dir, f'{img_id:06d}'+'.jpg'))
           
                # when saving only 1 img per species (debug) 
                # if id2name[int(class_id)].split(',')[-1] not in species_names:
                #     species_names.append(id2name[int(class_id)].split(',')[-1])
                #     os.makedirs(cropped_img_dir, exist_ok=True)
                #     img_crop.save(os.path.join(cropped_img_dir, f'{img_id:06d}'+'.jpg'))
                #     save_img_num += 1

                img_id += 1
                if img_id % 5000 == 0:
                    print('progress:',img_id)


        # get collection point and date from taskinfo-file using taskname
        task_name = (task_dir.split('/')[-1])[5:-29] # taskname without time stamp
        wb = openpyxl.load_workbook(os.path.join(root_dir, 'cvat_taskinfo.xlsx')) # path to cvat taskinfo
        sheet = wb['x' + args.mag] 
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[4] is not None and task_name == str(row[4]).strip().lower():
                point_ja = row[5]
                date = row[6]
                point = row[10]

        # add data to anno_list
        with open(save_dir_path + '/anno_list_' + args.mag + '.csv', mode='a',newline='',encoding='utf-8') as f:
            writer = csv.writer(f)
            for i in range(len(path_data)):
                data = id_data[i] + label_data[i] + subsp_data[i] + remark_data[i] + path_data[i] + crop_data[i] + [point] + [point_ja] + [date] + [task_name]
                writer.writerow(data)


# output dataset info 
print('crowd_imgs:',crowd_img_count)
print('cropped_imgs:',img_id)
print('tasks:',len(task_zip_paths))


        
            